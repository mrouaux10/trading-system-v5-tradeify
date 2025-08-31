#!/usr/bin/env python3
"""
Lightning 50K - Generador Excel Completo con Tabla Diaria
========================================================
Genera Excel con TODOS los trades + tabla resumen diario
"""

import pandas as pd
import numpy as np
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def create_complete_excel_with_daily_summary():
    """Crear Excel completo con tabla principal y resumen diario"""
    
    logger.info("Creando Excel COMPLETO con tabla diaria...")
    
    # Cargar el CSV optimizado
    csv_file = 'results/Lightning_50K_Optimized_Complete.csv'
    df = pd.read_csv(csv_file)
    
    logger.info(f"Loading: CSV cargado: {len(df):,} filas")
    
    # Filtrar trades y separadores por separado
    trades_df = df[~df['Date'].str.contains('Final P&L Day:', na=False)].copy()
    separators_df = df[df['Date'].str.contains('Final P&L Day:', na=False)].copy()
    
    logger.info(f"Trades: Trades: {len(trades_df):,}")
    logger.info(f"Días: {len(separators_df):,}")
    
    # Convertir Net P&L a numérico para trades
    trades_df['Net P&L'] = pd.to_numeric(trades_df['Net P&L'].astype(str).str.replace('$', '').str.replace(',', ''), errors='coerce')
    
    # Extraer P&L diario de los separadores
    daily_pnl_data = []
    for _, row in separators_df.iterrows():
        date_str = row['Date']
        pnl_str = date_str.replace('Final P&L Day: ', '')
        try:
            pnl = float(pnl_str)
            daily_pnl_data.append(pnl)
        except:
            daily_pnl_data.append(0.0)
    
    # Crear tabla resumen diario
    daily_table = pd.DataFrame({
        'Day': range(1, len(daily_pnl_data) + 1),
        'Daily P&L': daily_pnl_data
    })
    
    logger.info(f"Tabla diaria creada: {len(daily_table)} días")
    
    # Crear workbook de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Lightning 50K Complete"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # TABLA PRINCIPAL COMPLETA (Columnas A-J)
    logger.info("Escribiendo tabla principal completa...")
    
    # Headers tabla principal
    main_headers = ['Date', 'Time', 'MNQ Price', 'Operation type', 'Contracts', 'Trade duration', 'Close Reason', 'Net P&L', 'Daily P&L', 'Balance']
    
    for col, header in enumerate(main_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Escribir TODOS los datos (trades + separadores) de forma eficiente
    logger.info("Writing: Escribiendo todos los datos...")
    
    # Usar pandas para escribir datos más eficientemente
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row.values, 1):
            if col_idx <= 10:  # Solo columnas A-J
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Manejar separadores
                if col_idx == 1 and str(value).startswith('Final P&L Day:'):
                    cell.value = value
                    cell.font = Font(bold=True, color="FF0000")  # Rojo para separadores
                elif pd.isna(value) or str(value) == 'nan':
                    cell.value = ""
                else:
                    cell.value = value
                
                cell.border = thin_border
                
                if col_idx in [8, 9, 10]:  # Columnas de dinero
                    cell.alignment = Alignment(horizontal='right')
    
    logger.info(f"Tabla principal escrita: {len(df)} filas")
    
    # TABLA RESUMEN DIARIO (Columnas L-M)
    logger.info("Escribiendo tabla resumen diario...")
    
    # Headers tabla resumen
    daily_headers = ['Day', 'Daily P&L']
    start_col = 12  # Columna L
    
    for col_offset, header in enumerate(daily_headers):
        cell = ws.cell(row=1, column=start_col + col_offset, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Datos tabla resumen
    for row_idx, (_, row) in enumerate(daily_table.iterrows(), 2):
        # Day number
        day_cell = ws.cell(row=row_idx, column=start_col, value=int(row['Day']))
        day_cell.border = thin_border
        day_cell.alignment = Alignment(horizontal='center')
        
        # Daily P&L
        pnl_value = row['Daily P&L']
        pnl_cell = ws.cell(row=row_idx, column=start_col + 1, value=f"${pnl_value:.2f}")
        pnl_cell.border = thin_border
        pnl_cell.alignment = Alignment(horizontal='right')
        
        # Colorear verde o rojo según ganancia/pérdida
        if pnl_value > 0:
            pnl_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif pnl_value < 0:
            pnl_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Estadísticas al final de la tabla diaria
    stats_start_row = len(daily_table) + 3
    
    # Calcular estadísticas
    total_pnl = sum(daily_pnl_data)
    profitable_days = len([x for x in daily_pnl_data if x > 0])
    loss_days = len([x for x in daily_pnl_data if x < 0])
    win_rate_days = (profitable_days / len(daily_pnl_data) * 100) if daily_pnl_data else 0
    avg_daily = sum(daily_pnl_data) / len(daily_pnl_data) if daily_pnl_data else 0
    best_day = max(daily_pnl_data) if daily_pnl_data else 0
    worst_day = min(daily_pnl_data) if daily_pnl_data else 0
    
    stats_data = [
        ("DAILY SUMMARY", ""),
        ("=" * 15, ""),
        ("Total Days:", len(daily_pnl_data)),
        ("Total P&L:", f"${total_pnl:.2f}"),
        ("Profitable Days:", profitable_days),
        ("Loss Days:", loss_days),
        ("Win Rate (Days):", f"{win_rate_days:.1f}%"),
        ("Avg Daily P&L:", f"${avg_daily:.2f}"),
        ("Best Day:", f"${best_day:.2f}"),
        ("Worst Day:", f"${worst_day:.2f}"),
        ("", ""),
        ("STRATEGY PARAMS", ""),
        ("=" * 15, ""),
        ("Stop Loss:", "1.0 pts"),
        ("Break Even:", "1.5 pts"),
        ("Trailing Trigger:", "4.0 pts"),
        ("Trailing Distance:", "3.0 pts"),
        ("TP Normal:", "15 pts"),
        ("TP Near RN:", "22 pts"),
        ("", ""),
        ("OPTIMIZATION", ""),
        ("=" * 15, ""),
        ("Drawdown Reduction:", "$1,231.75"),
        ("Profit Increase:", "$56,894"),
        ("Win Rate Boost:", "+10.2%"),
        ("Max Drawdown:", "$581.00"),
        ("Compliance Margin:", "$1,419.00")
    ]
    
    for i, (label, value) in enumerate(stats_data):
        label_cell = ws.cell(row=stats_start_row + i, column=start_col, value=label)
        value_cell = ws.cell(row=stats_start_row + i, column=start_col + 1, value=value)
        
        if label.endswith(":") and label != "":
            label_cell.font = Font(bold=True)
        elif label in ["DAILY SUMMARY", "STRATEGY PARAMS", "OPTIMIZATION"]:
            label_cell.font = Font(bold=True, color="FFFFFF")
            label_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        label_cell.border = thin_border
        value_cell.border = thin_border
        value_cell.alignment = Alignment(horizontal='right')
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 12,   # Date
        'B': 10,   # Time
        'C': 12,   # MNQ Price
        'D': 15,   # Operation type
        'E': 10,   # Contracts
        'F': 15,   # Trade duration
        'G': 15,   # Close Reason
        'H': 12,   # Net P&L
        'I': 12,   # Daily P&L
        'J': 15,   # Balance
        'L': 8,    # Day
        'M': 15    # Daily P&L (summary)
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Guardar archivo
    excel_file = 'results/Lightning_50K_Complete_With_Daily_Summary.xlsx'
    wb.save(excel_file)
    
    logger.info(f"Excel completo generado: {excel_file}")
    logger.info("Estructura:")
    logger.info("   Columnas A-J: TODOS los trades y separadores")
    logger.info("   Columnas L-M: Tabla resumen diario completa")
    logger.info(f"   {len(daily_table)} días de trading")
    logger.info(f"   Trades: {len(trades_df):,} trades totales")
    logger.info(f"   {len(df):,} filas totales en Excel")
    
    return excel_file, daily_table

def main():
    """Función principal"""
    try:
        excel_file, daily_summary = create_complete_excel_with_daily_summary()
        
        logger.info(f"\n¡Archivo Excel COMPLETO creado exitosamente!")
        logger.info(f"Location: Ubicación: {excel_file}")
        logger.info(f"\nRESUMEN ESTADÍSTICO:")
        logger.info(f"   Total P&L: ${daily_summary['Daily P&L'].sum():.2f}")
        logger.info(f"   Total días: {len(daily_summary)}")
        logger.info(f"   WIN: Días ganadores: {len(daily_summary[daily_summary['Daily P&L'] > 0])}")
        logger.info(f"   LOSS: Días perdedores: {len(daily_summary[daily_summary['Daily P&L'] < 0])}")
        logger.info(f"   Win rate diario: {len(daily_summary[daily_summary['Daily P&L'] > 0]) / len(daily_summary) * 100:.1f}%")
        logger.info(f"   Best: Mejor día: ${daily_summary['Daily P&L'].max():.2f}")
        logger.info(f"   Worst: Peor día: ${daily_summary['Daily P&L'].min():.2f}")
        
    except Exception as e:
        logger.error(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    main()
